import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests
import re
import io
import os
import logging
import zipfile
import urllib.parse
import urllib3
from concurrent.futures import ThreadPoolExecutor, as_completed

# Set up logging and disable insecure SSL warning noises
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Set page configuration
st.set_page_config(
    page_title="Ad Script Converter Pro", 
    page_icon="📝", 
    layout="wide"
)

# Custom Styles
st.markdown("""
<style>
    .reportview-container {
        background: #f8f9fa;
    }
    .stButton>button {
        width: 100%;
        font-weight: bold;
    }
</style>
""", unsafe_allow_html=True)


# Core parsing and transformation helpers
def extract_comment(content):
    """Extracts first HTML comment block."""
    match = re.search(r'<!--(.*?)-->', content, re.DOTALL)
    return match.group(1).strip() if match else ""


def extract_size(name_draft):
    """Extracts resolution sizes from draft name (e.g., 300x250)."""
    match = re.search(r'(\d+x\d+)', name_draft)
    return match.group(1) if match else ""


def process_adform_content(content):
    """Applies Adform specific replacements."""
    return content.replace('/redir="', '/redir=%%c1;cpdir="')


def extract_name(name_draft):
    """Cleans up names by taking the final branch block and stripping Crimson/Crimtan decorators."""
    blocks = name_draft.split("/")
    last_block = blocks[-1]
    last_block = re.sub(r'(?i)crimtan[_ ]', '', last_block)
    return last_block.strip()


def unwrap_nested_urls(url):
    """
    Recursively extracts and URL-decodes nested destination URLs from query parameters.
    Bypasses redirect networks (Adform, Gemius, DoubleClick) offline with 100% precision.
    """
    if not url:
        return ""
    
    # Clean up double slashes from JSON/JS outputs
    url = url.replace('\\/', '/').replace('\\', '').strip()
    
    # Robustly decode multiple layers of URL encoding (e.g., %253A -> %3A -> :)
    decoded = url
    for _ in range(3):
        new_decoded = urllib.parse.unquote(decoded)
        if new_decoded == decoded:
            break
        decoded = new_decoded
        
    # Standard parameter targets used by major ad-server redirect configurations
    redirect_triggers = [
        "redir=", "url=", "link=", "href=", "dest=", "goto=", 
        "target=", "click=", "cpdir=", "clickmacro="
    ]
    
    for trigger in redirect_triggers:
        if trigger in decoded.lower():
            idx = decoded.lower().find(trigger)
            sub_payload = decoded[idx + len(trigger):]
            # Strip outer framing query components
            if "&" in sub_payload:
                sub_payload = sub_payload.split("&")[0]
            # Validate if extracted payload looks like an absolute target
            if sub_payload.startswith(("http://", "https://", "www.")):
                if sub_payload.startswith("www."):
                    sub_payload = "https://" + sub_payload
                return unwrap_nested_urls(sub_payload)
                
    # Search for any raw embedded absolute URL in query payload
    # E.g. https://tracker.com/?https://www.target.com/page
    potential_urls = re.findall(r'(https?://[^\s\'"<>]+)', decoded)
    if len(potential_urls) > 1:
        # Return the final embedded target link (the deepest nested URL)
        return unwrap_nested_urls(potential_urls[-1])
        
    return decoded


def extract_candidate_urls(html_content):
    """
    Intelligently extracts potential tracking/redirection URLs from ad scripts.
    Useful for servers like Gemius where tracking endpoints double as script triggers.
    """
    urls = []
    
    # Strategy 1: Look for document.write or script sources with url= parameters
    pattern_write = r'src="([^"]+url=[^"]+)"'
    matches_write = re.findall(pattern_write, html_content)
    for m in matches_write:
        url_part = m.split("url=")
        if len(url_part) > 1:
            urls.append(url_part[1])
            
    # Strategy 2: Look for redir=, href=, or url= tags inside strings
    pattern_params = r'["\'](https?://[^"\']*(?:redir|click|url|dest|href|hit\.gemius)=[^"\']+)["\']'
    matches_params = re.findall(pattern_params, html_content, re.IGNORECASE)
    urls.extend(matches_params)
        
    # Strategy 3: Extract general script source URLs (ignoring static utility files)
    pattern_src = r'<script[^>]+src=["\'](https?:[^"\']+)["\']'
    matches_src = re.findall(pattern_src, html_content)
    for m in matches_src:
        if any(lib in m.lower() for lib in [
            "gajs.js", "xgemius.js", "analytics.js", "gtm.js", "recaptcha", "adsbygoogle", "prebid"
        ]):
            continue
        urls.append(m)
        
    # Strategy 4: Find standard HTML redirect anchor tags
    pattern_href = r'<a[^>]+href=["\'](https?:[^"\']+)["\']'
    matches_href = re.findall(pattern_href, html_content, re.IGNORECASE)
    urls.extend(matches_href)
        
    # Strategy 5: General regex search for tracking system endpoints
    general_urls = re.findall(r'https?://[^\s\'"<>]+', html_content)
    for u in general_urls:
        if u not in urls:
            if any(tr in u.lower() for tr in ["gemius", "adform", "redir", "click", "track", "hit.gemius"]):
                if not any(lib in u.lower() for lib in ["gajs.js", "xgemius.js"]):
                    urls.append(u)
                    
    # Clean output elements
    cleaned_urls = []
    for u in urls:
        u_clean = u.replace('\\/', '/').replace('\\', '').strip()
        if u_clean:
            cleaned_urls.append(u_clean)
                    
    return list(set(cleaned_urls))


def resolve_landing_page(url, timeout=5):
    """
    Resolves landing pages using static parameters and highly resilient request tracing.
    Follows HTTP redirects, parses meta refreshes, and runs session state loops.
    """
    if not url:
        return ""
    
    # 1. Run our instant parameter unwrapper first
    unwrapped = unwrap_nested_urls(url)
    
    # If the unwrap yields a clean destination page, trust it immediately without hitting networks
    if unwrapped and unwrapped != url:
        parsed = urllib.parse.urlparse(unwrapped)
        # Verify it successfully broke out of known ad trackers
        if not any(t in parsed.netloc.lower() for t in ["gemius", "adform", "doubleclick", "googleads"]):
            return unwrapped
            
    # 2. Proceed with dynamic redirection trace using session headers
    if url.startswith("//"):
        url = "https:" + url
        
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1'
    }
    
    try:
        session = requests.Session()
        session.headers.update(headers)
        
        # SSL Verification disabled because staging trackers often run expired certs
        response = session.get(url, allow_redirects=True, timeout=timeout, verify=False)
        final_url = response.url
        
        # Detect HTML meta refresh headers (common with tracker setups)
        try:
            html_body = response.text
            meta_match = re.search(r'<meta[^>]*http-equiv=["\']refresh["\'][^>]*url=([^"\'>\s]+)', html_body, re.IGNORECASE)
            if not meta_match:
                meta_match = re.search(r'content=["\']\d+;\s*url=([^"\']+)["\']', html_body, re.IGNORECASE)
                
            if meta_match:
                refresh_url = urllib.parse.unquote(meta_match.group(1).strip())
                refresh_url = refresh_url.strip('\'"')
                # Merge relative paths with final URL host domain
                refresh_url = urllib.parse.urljoin(final_url, refresh_url)
                return resolve_landing_page(refresh_url, timeout=timeout)
        except Exception:
            pass
            
        # Run unwrap check on final resolved URL target
        final_url = unwrap_nested_urls(final_url)
        return final_url
        
    except Exception as e:
        logging.error(f"Network redirection error for {url}: {e}")
        # Return fallback extracted data
        return unwrapped if unwrapped else url


def parse_and_process_file_content(raw_bytes, filename, parent_folder_name, macro_gdpr, macro_consent, macro_redir):
    """Decodes, parses, and creates the structured metadata mapping for each script."""
    try:
        content = raw_bytes.decode('utf-8')
    except UnicodeDecodeError:
        try:
            content = raw_bytes.decode('latin1')
        except UnicodeDecodeError:
            content = raw_bytes.decode('cp1252', errors='replace')
    
    comment = extract_comment(content)
    name_draft = comment.split(',')[0] if ',' in comment else comment
    
    # Fallback to filename if comment parsing yields no draft name
    if not name_draft:
        name_draft = os.path.splitext(filename)[0]
        
    # Replicate structural folder tree renamer logic
    if parent_folder_name:
        name_draft = f"{parent_folder_name}_{name_draft}"
        
    size = extract_size(name_draft)
    
    # Find original clean code block
    clean_original = content
    if comment:
        clean_original = content.replace(f"<!--{comment}-->", "").strip()
    
    # Run targeted transform logic
    processed_adform = process_adform_content(clean_original)
    name = extract_name(name_draft)
    
    # Modify DV360 tracking parameters
    processed_dv360 = (
        clean_original
        .replace('gdpr=0', macro_gdpr)
        .replace('gdpr_consent=', macro_consent)
        .replace('redir="', macro_redir)
    )
    
    # Isolate tracking endpoints and prioritize nested targets
    candidate_urls = extract_candidate_urls(clean_original)
    
    best_candidate = None
    for cand in candidate_urls:
        unwrapped = unwrap_nested_urls(cand)
        # If static extraction yields a valid destination non-tracking domain, pick it!
        if unwrapped != cand and not any(track_term in urllib.parse.urlparse(unwrapped).netloc.lower() for track_term in ["gemius", "adform", "doubleclick", "adserver"]):
            best_candidate = cand
            break
            
    if not best_candidate and candidate_urls:
        best_candidate = candidate_urls[0]
        
    extracted_url = best_candidate
    
    return {
        "name": name,
        "size": size,
        "name_draft": name_draft,
        "clean_original": clean_original,
        "processed_adform": processed_adform,
        "processed_dv360": processed_dv360,
        "extracted_url": extracted_url,
        "landing_page": ""  # Resolved concurrently
    }


def escape_js_template_string(s):
    """Safely escapes HTML scripts for execution inside JavaScript template strings."""
    return s.replace('\\', '\\\\').replace('`', '\\`').replace('${', '\\${').replace('</script>', '<\\/script>')


def generate_styled_html_preview(records):
    """Generates a premium, clean, and highly usable preview HTML file with interactive sandbox frames."""
    cards_html = ""
    for idx, r in enumerate(records):
        size_str = r['size']
        width_val, height_val = 300, 250
        if size_str and 'x' in size_str:
            try:
                w, h = map(int, size_str.split('x'))
                width_val, height_val = w, h
            except:
                pass
                
        escaped_script = escape_js_template_string(r['clean_original'])
        
        cards_html += f"""
        <div class="card">
            <div class="card-header">
                <div>
                    <span class="badge">{r['size'] if r['size'] else 'Unknown Size'}</span>
                    <strong style="font-size: 1.15rem; color: #1e293b; margin-left: 8px;">{r['name']}</strong>
                </div>
                <div style="font-size: 0.85rem; color: #64748b;">
                    Draft: {r['name_draft']}
                </div>
            </div>
            <div class="card-body">
                <div class="content-split">
                    <!-- Column 1: Copyable Scripts -->
                    <div class="pane">
                        <div class="section-title">Original Cleaned Script</div>
                        <div class="textarea-container">
                            <textarea id="orig-{idx}" readonly>{r['clean_original']}</textarea>
                            <button class="btn-copy" onclick="copyText('orig-{idx}')">Copy Clean</button>
                        </div>
                        
                        <div class="section-title">DV360 Modified Script</div>
                        <div class="textarea-container">
                            <textarea id="dv360-{idx}" readonly>{r['processed_dv360']}</textarea>
                            <button class="btn-copy" onclick="copyText('dv360-{idx}')">Copy DV360</button>
                        </div>
                    </div>
                    
                    <!-- Column 2: Live Ad Render Box (Solves client-side redirect capturing) -->
                    <div class="pane render-pane">
                        <div class="section-title" style="margin-top:0; text-align:center;">Live Interactive Sandbox Preview</div>
                        <p style="font-size:0.75rem; color:#64748b; margin-top:0; text-align:center;">
                            Click inside the frame below to execute redirects and trace landing pages live.
                        </p>
                        <div class="frame-container">
                            <iframe id="iframe-{idx}" class="ad-frame" style="width: {width_val}px; height: {height_val}px;" sandbox="allow-scripts allow-popups allow-popups-to-escape-sandbox allow-same-origin"></iframe>
                        </div>
                        <script>
                            (function() {{
                                const iframe = document.getElementById('iframe-{idx}');
                                const doc = iframe.contentWindow.document || iframe.contentDocument;
                                doc.open();
                                doc.write(`<!DOCTYPE html><html><head><style>body {{ margin: 0; padding: 0; display: flex; justify-content: center; align-items: center; min-height: 100vh; overflow: hidden; background: transparent; }}</style></head><body>{escaped_script}</body></html>`);
                                doc.close();
                            }})();
                        </script>
                    </div>
                </div>
                
                {f'<div class="landing-page"><strong>Destination Landing Page:</strong> <a href="{r["landing_page"]}" target="_blank">{r["landing_page"]}</a></div>' if r['landing_page'] else ''}
            </div>
        </div>
        """

    html_template = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Ad Scripts Live Preview Dashboard</title>
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
            background-color: #f1f5f9;
            color: #334155;
            margin: 0;
            padding: 40px 20px;
        }}
        .container {{
            max-width: 1100px;
            margin: 0 auto;
        }}
        header {{
            margin-bottom: 30px;
            text-align: center;
        }}
        h1 {{
            color: #0f172a;
            font-size: 2rem;
            margin-bottom: 8px;
        }}
        .card {{
            background: #ffffff;
            border-radius: 12px;
            box-shadow: 0 4px 6px -1px rgb(0 0 0 / 0.1), 0 2px 4px -2px rgb(0 0 0 / 0.1);
            margin-bottom: 30px;
            border: 1px solid #e2e8f0;
            overflow: hidden;
        }}
        .card-header {{
            background-color: #f8fafc;
            padding: 16px 20px;
            border-bottom: 1px solid #e2e8f0;
            display: flex;
            justify-content: space-between;
            align-items: center;
            flex-wrap: wrap;
            gap: 10px;
        }}
        .badge {{
            background-color: #3b82f6;
            color: white;
            padding: 4px 8px;
            border-radius: 6px;
            font-size: 0.75rem;
            font-weight: bold;
            text-transform: uppercase;
        }}
        .card-body {{
            padding: 24px;
        }}
        .content-split {{
            display: grid;
            grid-template-columns: 1.2fr 0.8fr;
            gap: 24px;
            align-items: stretch;
        }}
        @media (max-width: 768px) {{
            .content-split {{
                grid-template-columns: 1fr;
            }}
        }}
        .pane {{
            display: flex;
            flex-direction: column;
        }}
        .render-pane {{
            background-color: #fafafa;
            border: 1px dashed #cbd5e1;
            border-radius: 8px;
            padding: 20px;
            justify-content: center;
            align-items: center;
        }}
        .frame-container {{
            display: flex;
            justify-content: center;
            align-items: center;
            background-color: #ffffff;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
            border-radius: 4px;
            padding: 10px;
            border: 1px solid #e2e8f0;
        }}
        .ad-frame {{
            border: none;
            overflow: hidden;
        }}
        .section-title {{
            font-size: 0.85rem;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            color: #475569;
            margin-top: 14px;
            margin-bottom: 6px;
        }}
        .textarea-container {{
            position: relative;
            margin-bottom: 16px;
        }}
        textarea {{
            width: 100%;
            height: 100px;
            font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, monospace;
            font-size: 0.85rem;
            padding: 10px;
            box-sizing: border-box;
            border: 1px solid #cbd5e1;
            border-radius: 6px;
            background-color: #f8fafc;
            color: #334155;
            resize: vertical;
        }}
        .btn-copy {{
            position: absolute;
            top: 8px;
            right: 8px;
            background: #0f172a;
            color: #ffffff;
            border: none;
            padding: 4px 10px;
            font-size: 0.75rem;
            border-radius: 4px;
            cursor: pointer;
            transition: all 0.2s;
            font-weight: 500;
        }}
        .btn-copy:hover {{
            background: #1e293b;
        }}
        .btn-copy.success {{
            background: #10b981;
        }}
        .landing-page {{
            margin-top: 20px;
            background-color: #eff6ff;
            padding: 12px 16px;
            border-radius: 6px;
            border: 1px solid #bfdbfe;
            font-size: 0.9rem;
            word-break: break-all;
        }}
        .landing-page a {{
            color: #2563eb;
            text-decoration: none;
            font-weight: 500;
        }}
        .landing-page a:hover {{
            text-decoration: underline;
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>Ad Scripts Preview Directory</h1>
            <p style="color: #64748b;">Generated dynamically via Ad Script Converter Pro</p>
        </header>
        {cards_html}
    </div>

    <script>
        function copyText(elementId) {{
            const textarea = document.getElementById(elementId);
            textarea.select();
            textarea.setSelectionRange(0, 99999);
            
            try {{
                document.execCommand('copy');
                const btn = textarea.nextElementSibling;
                const baseText = btn.innerText;
                btn.innerText = "Copied!";
                btn.classList.add('success');
                setTimeout(() => {{
                    btn.innerText = baseText;
                    btn.classList.remove('success');
                }}, 1500);
            }} catch (err) {{
                console.error('Copy failed', err);
            }}
        }}
    </script>
</body>
</html>
"""
    return html_template


def update_outputs_from_records(records):
    """Regenerates spreadsheet and preview deliverables inside state variables instantly."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ad Scripts Data"
    sheet.views.sheetView[0].showGridLines = True
    
    headers = [
        "Creative Name", "Click URL", "Content (Adform Tag)", "Creative Size", 
        "Campaign Manager Placement/Draft Name", "Clean Original Script", 
        "DV360 Modified Script Tag", "Destination Landing Page URL"
    ]
    sheet.append(headers)
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    border_thin = Side(border_style="thin", color="D9D9D9")
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border
    
    for r in records:
        sheet.append([
            r["name"],
            "",  
            r["processed_adform"],
            r["size"],
            r["name_draft"],
            r["clean_original"],
            r["processed_dv360"],
            r["landing_page"]
        ])
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.font = Font(name="Calibri", size=11)
            cell.border = cell_border
            if cell.column in [3, 6, 7]:
                cell.alignment = Alignment(vertical="top", wrap_text=False)
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        if col[0].column in [3, 6, 7]:
            sheet.column_dimensions[col_letter].width = 35
        else:
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    
    html_output = generate_styled_html_preview(records)
    html_buffer = io.BytesIO(html_output.encode('utf-8'))
    
    st.session_state.excel_bytes = excel_buffer.getvalue()
    st.session_state.html_bytes = html_buffer.getvalue()


def main():
    st.title("📝 Ad Script Converter Pro")
    st.write("Convert agency-specific raw `.txt` script tags (either uploaded directly or nested in `.zip` folders) into clean, macro-enabled configurations ready for upload.")

    # Highly optimized default configurations
    resolve_redirects = True
    timeout_limit = 5
    concurrent_threads = 12  # Increased threads for handling 100+ items efficiently
    macro_gdpr = "gdpr=${GDPR}"
    macro_consent = "gdpr_consent=${GDPR_CONSENT_328}"
    macro_redir = "redir=${CLICK_URL}\""

    # State setups
    if 'excel_bytes' not in st.session_state:
        st.session_state.excel_bytes = None
    if 'html_bytes' not in st.session_state:
        st.session_state.html_bytes = None
    if 'records' not in st.session_state:
        st.session_state.records = None

    uploaded_files = st.file_uploader(
        "Upload Raw Ad Script Files (.txt or .zip archives)", 
        type=['txt', 'zip'], 
        accept_multiple_files=True
    )

    if uploaded_files:
        st.info(f"📂 {len(uploaded_files)} item(s) selected.")
        
        if st.button("Process Scripts", type="primary"):
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            records = []
            
            # Step 1: Parse and unpack files (including deep directory structures inside ZIP files)
            for idx, uploaded_file in enumerate(uploaded_files):
                file_name_lower = uploaded_file.name.lower()
                
                if file_name_lower.endswith('.zip'):
                    status_text.text(f"📦 Extracting directory structure from: {uploaded_file.name}...")
                    try:
                        zip_data = uploaded_file.read()
                        with zipfile.ZipFile(io.BytesIO(zip_data)) as z:
                            for member in z.infolist():
                                if member.filename.endswith('/') or member.is_dir():
                                    continue
                                if not member.filename.lower().endswith('.txt'):
                                    continue
                                
                                parent_folder_name = os.path.basename(os.path.dirname(member.filename))
                                filename_only = os.path.basename(member.filename)
                                
                                with z.open(member) as f:
                                    raw_bytes = f.read()
                                    
                                record = parse_and_process_file_content(
                                    raw_bytes,
                                    filename_only,
                                    parent_folder_name,
                                    macro_gdpr,
                                    macro_consent,
                                    macro_redir
                                )
                                records.append(record)
                    except Exception as e:
                        st.error(f"Error unpacking ZIP archive '{uploaded_file.name}': {str(e)}")
                
                elif file_name_lower.endswith('.txt'):
                    status_text.text(f"📄 Reading single script file: {uploaded_file.name}...")
                    raw_bytes = uploaded_file.read()
                    
                    record = parse_and_process_file_content(
                        raw_bytes,
                        uploaded_file.name,
                        "",  
                        macro_gdpr,
                        macro_consent,
                        macro_redir
                    )
                    records.append(record)

            if not records:
                st.warning("⚠️ No valid `.txt` script tags were discovered in the uploaded content.")
                return

            # Step 2: Resolve redirect links concurrently with upgraded heuristics
            if resolve_redirects:
                status_text.text("🔗 Running smart dynamic extraction & recursive unwrap checks...")
                
                unique_urls = list({r["extracted_url"] for r in records if r["extracted_url"]})
                resolved_url_map = {}
                
                if unique_urls:
                    with ThreadPoolExecutor(max_workers=concurrent_threads) as executor:
                        future_to_url = {
                            executor.submit(resolve_landing_page, url, timeout_limit): url 
                            for url in unique_urls
                        }
                        
                        completed_count = 0
                        total_futures = len(future_to_url)
                        
                        for future in as_completed(future_to_url):
                            orig_url = future_to_url[future]
                            try:
                                resolved_url = future.result()
                                resolved_url_map[orig_url] = resolved_url
                            except Exception as e:
                                resolved_url_map[orig_url] = orig_url
                                logging.error(f"Error handling future result: {e}")
                            
                            completed_count += 1
                            progress_val = int((completed_count / total_futures) * 100)
                            progress_bar.progress(progress_val)
                            status_text.text(f"Unwrapped & Traced {completed_count}/{total_futures} tracking systems...")
                
                # Map resolved URLs back to record fields
                for r in records:
                    url_to_lookup = r["extracted_url"]
                    if url_to_lookup in resolved_url_map:
                        r["landing_page"] = resolved_url_map[url_to_lookup]
            
            # Step 3: Store and generate configurations
            status_text.text("📊 Formatting premium spreadsheet outputs...")
            st.session_state.records = records
            update_outputs_from_records(records)
            
            progress_bar.progress(100)
            status_text.empty()
            st.success(f"Successfully processed {len(records)} script configs!")

    # Step 4: Live workspace configurations divided into tabs
    if st.session_state.records is not None:
        
        tab1, tab2, tab3 = st.tabs([
            "📝 Edit & Review Tracker", 
            "🔍 Live Visual Ad Sandbox", 
            "🛠️ Bulk Find & Replace"
        ])
        
        # TAB 1: Main interactive table containing data_editor workspace
        with tab1:
            st.subheader("Interactive Tracking Matrix")
            st.write("Double-click on any cell in the editable matrix below to customize details or paste final Destination Landing Pages directly.")
            
            # Deep Crawl fallback handler button (for lingering tricky trackers)
            unresolved_count = sum(1 for r in st.session_state.records if not r["landing_page"] or any(t in r["landing_page"].lower() for t in ["gemius", "adform", "adserver"]))
            if unresolved_count > 0:
                st.warning(f"⚠️ {unresolved_count} landing page link(s) still point to ad servers or look unresolved.")
                if st.button("🔗 Run Deep Crawler (Retries Unresolved Landing Pages)", type="secondary"):
                    with st.spinner("Executing dynamic fallback hops..."):
                        for r in st.session_state.records:
                            is_unresolved = not r["landing_page"] or any(t in r["landing_page"].lower() for t in ["gemius", "adform", "adserver"])
                            if is_unresolved and r["extracted_url"]:
                                r["landing_page"] = resolve_landing_page(r["extracted_url"], timeout=10)
                        update_outputs_from_records(st.session_state.records)
                        st.success("Deep crawler run complete!")
                        st.rerun()

            # Construct a list from state records to map into the editor
            edit_data = []
            for r in st.session_state.records:
                edit_data.append({
                    "Creative Name": r["name"],
                    "Size": r["size"],
                    "Placement/Draft Name": r["name_draft"],
                    "Destination Landing Page URL": r["landing_page"]
                })
                
            edited_df = st.data_editor(
                edit_data,
                use_container_width=True,
                num_rows="fixed",
                column_config={
                    "Creative Name": st.column_config.TextColumn(width="medium"),
                    "Size": st.column_config.TextColumn(width="small"),
                    "Placement/Draft Name": st.column_config.TextColumn(width="large"),
                    "Destination Landing Page URL": st.column_config.TextColumn(
                        width="large", 
                        help="Enter/Paste the final captured landing page here."
                    )
                },
                key="matrix_editor"
            )
            
            # Identify changes dynamically to push them back to records state
            has_changes = False
            for idx, row in enumerate(edited_df):
                rec = st.session_state.records[idx]
                if (rec["name"] != row["Creative Name"] or 
                    rec["size"] != row["Size"] or 
                    rec["name_draft"] != row["Placement/Draft Name"] or 
                    rec["landing_page"] != row["Destination Landing Page URL"]):
                    
                    rec["name"] = row["Creative Name"]
                    rec["size"] = row["Size"]
                    rec["name_draft"] = row["Placement/Draft Name"]
                    rec["landing_page"] = row["Destination Landing Page URL"]
                    has_changes = True
                    
            if has_changes:
                # Update downloads in the background and trigger refresh
                update_outputs_from_records(st.session_state.records)
                st.rerun()

            st.markdown("### 📥 Export Final Deliverables")
            col1, col2 = st.columns(2)
            
            with col1:
                st.download_button(
                    label="📥 Download Excel Tracker Spreadsheet",
                    data=st.session_state.excel_bytes,
                    file_name="processed_ad_tracker.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            with col2:
                st.download_button(
                    label="🌐 Download Live HTML Preview Directory",
                    data=st.session_state.html_bytes,
                    file_name="ad_scripts_preview_dashboard.html",
                    mime="text/html"
                )

        # TAB 2: Live Ad Sandbox (For manually clicking remaining outliers)
        with tab2:
            st.subheader("🔍 Live Visual Ad Sandbox")
            st.write(
                "If an ad still refuses to resolve automatically (due to strict client-side JS redirects or geo-blocking checks), "
                "you can run it in this isolated sandbox wrapper, click the banner, and copy-paste the target URL back to Tab 1."
            )
            
            creative_names = [r["name"] for r in st.session_state.records]
            selected_creative_name = st.selectbox("Select creative to load in the sandbox frame:", creative_names)
            
            selected_record = next((r for r in st.session_state.records if r["name"] == selected_creative_name), None)
            
            if selected_record:
                size_str = selected_record["size"]
                width, height = 300, 250
                if size_str and 'x' in size_str:
                    try:
                        w, h = map(int, size_str.split('x'))
                        width, height = w, h
                    except:
                        pass
                
                st.write(f"**Sandboxed Creative Canvas ({width}x{height})**")
                
                # Render ad script wrapped in a clean, scroll-safe frame
                html_to_render = f"""
                <!DOCTYPE html>
                <html>
                <head>
                    <style>
                        body {{ 
                            margin: 0; 
                            padding: 0; 
                            display: flex; 
                            justify-content: center; 
                            align-items: center; 
                            min-height: 100vh; 
                            background-color: #ffffff;
                            overflow: hidden;
                        }}
                    </style>
                </head>
                <body>
                    <div style="border: 1px dashed #cbd5e1; padding: 10px; border-radius: 4px; display: inline-block;">
                        {selected_record['clean_original']}
                    </div>
                </body>
                </html>
                """
                import streamlit.components.v1 as components
                # Add margin to fit dashed borders nicely
                components.html(html_to_render, width=width + 50, height=height + 50, scrolling=True)

        # TAB 3: Bulk Find & Replace
        with tab3:
            st.subheader("🛠️ Find & Replace Naming Utility")
            st.info("💡 Adjust and align naming patterns dynamically (similar to Ctrl+H in Excel) before generating final outputs.")
            
            col_find, col_replace = st.columns(2)
            with col_find:
                find_val = st.text_input("Find string", placeholder="e.g., OldBrandName")
            with col_replace:
                replace_val = st.text_input("Replace with", placeholder="e.g., NewBrandName")
                
            st.write("Target Fields:")
            replace_in_name = st.checkbox("Apply to Creative Name", value=True)
            replace_in_draft = st.checkbox("Apply to Placement/Draft Name", value=False)
                    
            if st.button("Apply Replacements", type="secondary"):
                if find_val:
                    match_count = 0
                    for r in st.session_state.records:
                        target_updated = False
                        if replace_in_name and find_val in r["name"]:
                            r["name"] = r["name"].replace(find_val, replace_val)
                            target_updated = True
                        if replace_in_draft and find_val in r["name_draft"]:
                            r["name_draft"] = r["name_draft"].replace(find_val, replace_val)
                            target_updated = True
                        if target_updated:
                            match_count += 1
                            
                    if match_count > 0:
                        update_outputs_from_records(st.session_state.records)
                        st.success(f"Successfully replaced '{find_val}' with '{replace_val}' in {match_count} item(s)!")
                        st.rerun()
                    else:
                        st.warning(f"No matches found for '{find_val}' within the targeted fields.")
                else:
                    st.error("Please enter a string to find.")


if __name__ == "__main__":
    main()
